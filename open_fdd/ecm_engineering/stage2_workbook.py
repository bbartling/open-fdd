"""Stage-2 professional ECM workbook builder (Cover → Documentation).

Renders dual-rail Inputs with Assumption_Note as engineer-facing sheets.
Uses openpyxl when available; otherwise writes a minimal OOXML Inputs pack.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from .contracts import EngineeringInput, InputRail

STAGE2_SHEET_ORDER = (
    "Cover",
    "Calibrated_Twin",
    "Inputs",
    "Measure_Summary",
    "Equipment_Sizing",
    "Hours_Bases",
    "Package_Matrix",
    "Calculation_Trace",
    "Agent_Action_Log",
    "Publication_Gates",
    "Documentation",
)

INPUTS_COLUMNS = (
    "input_id",
    "display_name",
    "value",
    "unit",
    "rail",
    "source_type",
    "confidence",
    "editable",
    "validation_status",
    "assumption_method",
    "Assumption_Note",
    "linked_measure_ids",
    "source_reference",
)


def dual_rail_fixture_inputs() -> list[EngineeringInput]:
    """Synthetic Stage-1/2 fixture with at least one ss_* / ep_* sizing pair."""
    from .contracts import AssumptionMethod, SourceType

    return [
        EngineeringInput(
            input_id="ss_fan_hp",
            display_name="Spreadsheet fan power",
            value=25.0,
            unit="hp",
            rail=InputRail.SPREADSHEET,
            source_type=SourceType.NAMEPLATE,
            assumption_note="Nameplate from TAB report AHU-1 supply fan.",
            assumption_method=AssumptionMethod.NAMEPLATE,
            linked_measure_ids=["ECM-DSP-RESET", "ECM-FAN-SCHED"],
            source_reference="TAB-2024-AHU1",
        ),
        EngineeringInput(
            input_id="ep_fan_hp",
            display_name="EnergyPlus autosized fan power",
            value=22.4,
            unit="hp",
            rail=InputRail.ENERGYPLUS,
            source_type=SourceType.ENERGYPLUS_AUTOSIZED,
            assumption_note="From baseline .eio Fan:SystemModel Peak Design Power.",
            assumption_method=AssumptionMethod.EIO_AUTOSIZE,
            linked_measure_ids=["ECM-DSP-RESET"],
            source_reference="baseline.eio",
        ),
        EngineeringInput(
            input_id="ss_sched_hours_saved",
            display_name="Spreadsheet schedule hours saved",
            value=1200.0,
            unit="h/yr",
            rail=InputRail.SPREADSHEET,
            source_type=SourceType.AGENT_INFERRED,
            assumption_note=(
                "FLH back-calc from cascade kWh / ss_fan_kw — not Twin AMY calendar hours "
                "(BUG-ECM-014). Period = AMY calendar year matching Twin weather."
            ),
            assumption_method=AssumptionMethod.FLH_FROM_CASCADE,
            linked_measure_ids=["ECM-FAN-SCHED"],
        ),
        EngineeringInput(
            input_id="ep_sched_hours_saved",
            display_name="EnergyPlus schedule hours saved",
            value=1085.0,
            unit="h/yr",
            rail=InputRail.ENERGYPLUS,
            source_type=SourceType.ENERGYPLUS_DERIVED,
            assumption_note="FanAvail calendar delta baseline vs schedule measure (AMY).",
            assumption_method=AssumptionMethod.AMY_CALENDAR_HOURS,
            linked_measure_ids=["ECM-FAN-SCHED"],
            source_reference="cascade/FanAvail",
        ),
    ]


def build_stage2_workbook(
    output_path: str | Path,
    *,
    inputs: list[EngineeringInput] | None = None,
    project_name: str = "Open-FDD ECM Package",
    facility_name: str = "Synthetic Facility",
    action_log: list[dict[str, Any]] | None = None,
) -> Path:
    """Write a Stage-2 professional workbook with dual-rail Inputs + Assumption_Note."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    rows = inputs if inputs is not None else dual_rail_fixture_inputs()
    log = action_log or []

    try:
        from openpyxl import Workbook
        from openpyxl.workbook.defined_name import DefinedName
    except ImportError:
        return _build_minimal_inputs_json_sidecar(output, rows, project_name, facility_name, log)

    wb = Workbook()
    # Cover
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = project_name
    cover["A2"] = facility_name
    cover["A3"] = "Stage 2 professional ECM workbook (Open-FDD owned)"
    cover["A4"] = "Real EnergyPlus / IDF work remains in vibe20; this book holds calcs + Inputs."

    for name in STAGE2_SHEET_ORDER[1:]:
        if name not in wb.sheetnames:
            wb.create_sheet(name)

    twin = wb["Calibrated_Twin"]
    twin["A1"] = "Calibrated Twin (reference only)"
    twin["A2"] = "Populate from vibe20 ecm_simulation_evidence.json on import."

    inputs_ws = wb["Inputs"]
    for col, header in enumerate(INPUTS_COLUMNS, start=1):
        inputs_ws.cell(1, col, header)
    for r_i, inp in enumerate(rows, start=2):
        d = inp.as_dict()
        values = [
            d.get("input_id"),
            d.get("display_name"),
            d.get("value"),
            d.get("unit"),
            d.get("rail"),
            d.get("source_type"),
            d.get("confidence"),
            d.get("editable"),
            d.get("validation_status"),
            d.get("assumption_method"),
            d.get("assumption_note"),
            ",".join(d.get("linked_measure_ids") or []),
            d.get("source_reference"),
        ]
        for c_i, val in enumerate(values, start=1):
            inputs_ws.cell(r_i, c_i, val)

    # Stable named ranges for dual-rail columns (table-like, not cell B17).
    # Input_id column + Assumption_Note column whole ranges.
    n_rows = max(len(rows), 1) + 1
    wb.defined_names.add(
        DefinedName(name="Inputs_Table", attr_text=f"Inputs!$A$1:$M${n_rows}")
    )
    wb.defined_names.add(
        DefinedName(name="Assumption_Note", attr_text=f"Inputs!$K$2:$K${n_rows}")
    )
    wb.defined_names.add(DefinedName(name="Inputs_input_id", attr_text=f"Inputs!$A$2:$A${n_rows}"))

    eq = wb["Equipment_Sizing"]
    eq["A1"] = "input_id"
    eq["B1"] = "rail"
    eq["C1"] = "value"
    eq["D1"] = "unit"
    ri = 2
    for inp in rows:
        if "fan" in inp.input_id or "cooling" in inp.input_id or "ahu" in inp.input_id:
            eq.cell(ri, 1, inp.input_id)
            eq.cell(ri, 2, inp.rail.value)
            eq.cell(ri, 3, inp.value)
            eq.cell(ri, 4, inp.unit)
            ri += 1

    hours = wb["Hours_Bases"]
    hours["A1"] = "input_id"
    hours["B1"] = "rail"
    hours["C1"] = "value"
    hours["D1"] = "Assumption_Note"
    ri = 2
    for inp in rows:
        if "hour" in inp.input_id.lower():
            hours.cell(ri, 1, inp.input_id)
            hours.cell(ri, 2, inp.rail.value)
            hours.cell(ri, 3, inp.value)
            hours.cell(ri, 4, inp.assumption_note)
            ri += 1

    summary = wb["Measure_Summary"]
    summary["A1"] = "measure_id"
    summary["B1"] = "note"
    summary["A2"] = "(from evidence import)"

    pkg = wb["Package_Matrix"]
    pkg["A1"] = "Stage 5 enforces package/interaction matrix; Stage 2 reserves the sheet."

    trace = wb["Calculation_Trace"]
    trace["A1"] = "trace_id"
    trace["B1"] = "equation"
    trace["C1"] = "inputs_used"
    trace["D1"] = "result"

    log_ws = wb["Agent_Action_Log"]
    log_ws["A1"] = "action"
    log_ws["B1"] = "input_id"
    log_ws["C1"] = "reason"
    log_ws["D1"] = "assumption_note"
    for i, entry in enumerate(log, start=2):
        log_ws.cell(i, 1, entry.get("action"))
        log_ws.cell(i, 2, entry.get("input_id"))
        log_ws.cell(i, 3, entry.get("reason"))
        log_ws.cell(i, 4, entry.get("assumption_note"))

    gates = wb["Publication_Gates"]
    gates["A1"] = "gate"
    gates["B1"] = "status"
    gates["A2"] = "assumption_note_required"
    gates["B2"] = "WARNING (enforced Stage 5)"
    gates["A3"] = "dual_rail_sizing_pair"
    gates["B3"] = "preview"

    docs = wb["Documentation"]
    docs["A1"] = "Ownership"
    docs["A2"] = "Open-FDD owns schemas, equations, provenance, workbook/DOCX builders."
    docs["A3"] = "Vibe20 owns IDF/MCP/sim and ecm_simulation_evidence.json export."
    docs["A4"] = "Never silently paste Twin calendar hours over spreadsheet FLH (BUG-ECM-014)."

    wb.save(output)
    sidecar = output.with_suffix(".ecm_engineering_inputs.json")
    sidecar.write_text(
        json.dumps(
            {
                "schema": "ecm_engineering_inputs_v1",
                "project": project_name,
                "facility": facility_name,
                "inputs": [i.as_dict() for i in rows],
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    return output


def _build_minimal_inputs_json_sidecar(
    output: Path,
    rows: list[EngineeringInput],
    project_name: str,
    facility_name: str,
    log: list[dict[str, Any]],
) -> Path:
    """Fallback when openpyxl is missing — still emit Inputs contract + placeholder xlsx note."""
    manifest = output.with_suffix(".ecm_engineering_inputs.json")
    manifest.write_text(
        json.dumps(
            {
                "schema": "ecm_engineering_inputs_v1",
                "project": project_name,
                "facility": facility_name,
                "inputs": [i.as_dict() for i in rows],
                "action_log": log,
                "warning": "openpyxl not installed; xlsx not written",
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    output.write_text(
        "Stage-2 workbook requires openpyxl; see sibling .ecm_engineering_inputs.json\n",
        encoding="utf-8",
    )
    return output


__all__ = [
    "STAGE2_SHEET_ORDER",
    "INPUTS_COLUMNS",
    "dual_rail_fixture_inputs",
    "build_stage2_workbook",
]
