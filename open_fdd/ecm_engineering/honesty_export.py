"""BUG-OFDD-ECM-009 — honesty / twin compare workbook export."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from .honesty_status import (
    MeasureHonestyStatus,
    classify_measure_status,
    pct_diff_industry_vs_eplus,
    wiring_echo_pct,
)

HONESTY_SHEETS = (
    "Contents",
    "Model_Provenance",
    "Inputs",
    "Industry_Screening",
    "Measures",
)


def build_honesty_workbook(
    output_path: str | Path,
    *,
    twin_payload: dict[str, Any] | None = None,
    job_name: str = "ECM honesty",
) -> Path:
    """Write Contents / Provenance / Inputs / Industry_Screening / Measures (+ optional).

    Skips Cover, Formula_Trace, and Documentation dumps (fold into Contents).
    Emits Demand when ``twin_payload`` includes demand; Twin_Calibrate when attached.
    """
    from openpyxl import Workbook

    payload = twin_payload or {}
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # openpyxl starts with Sheet — rename to Contents
    ws = wb.active
    ws.title = "Contents"
    _write_contents(ws, job_name, payload)

    _write_provenance(wb.create_sheet("Model_Provenance"), payload.get("provenance") or {})
    _write_inputs(wb.create_sheet("Inputs"), payload.get("inputs") or [])
    _write_industry(wb.create_sheet("Industry_Screening"), payload)
    _write_measures(
        wb.create_sheet("Measures"),
        payload.get("measures") or [],
        baseline_kwh=payload.get("baseline_kwh"),
        elec_rate=float((payload.get("inputs_map") or {}).get("elec_usd_per_kwh") or 0.14),
    )

    demand = payload.get("demand")
    if demand:
        _write_demand(wb.create_sheet("Demand"), demand)

    twin_cal = payload.get("twin_calibrate")
    if twin_cal:
        _write_twin_calibrate(wb.create_sheet("Twin_Calibrate"), twin_cal)

    wb.save(out)
    return out


def _write_contents(ws, job_name: str, payload: dict[str, Any]) -> None:
    ws["A1"] = f"Open-FDD ECM honesty — {job_name}"
    ws["A2"] = "How to read this workbook"
    rows = [
        (4, "Sheet", "Purpose"),
        (5, "Model_Provenance", "IDF / EPW / G14 / cascade timestamps"),
        (6, "Inputs", "Named building inputs; yellow FLH = FITTED_FROM_EPLUS (not validation)"),
        (7, "Industry_Screening", "Independent ESCO methods (2nd set of eyes)"),
        (8, "Measures", "Primary roster + Savings_Check; status FITTED/BALLPARK/NO_EP/FAIL_SIGN"),
        (9, "Demand", "July weekday / weekend / load-shed kW when twin payload present"),
        (10, "Twin_Calibrate", "Optional monthly twin calibration when attached"),
        (12, "Status meanings", ""),
        (13, "FITTED", "Hours reverse-solved from E+ so sheet≈E+; wiring_echo_pct≈0 is audit only"),
        (14, "BALLPARK", "Independent screening assumptions; use pct_diff_industry_vs_eplus"),
        (15, "NO_EP", "No EnergyPlus / patch missing — do not invent ep_kwh"),
        (16, "FAIL_SIGN", "Sheet vs E+ opposite sign (broken physics / patch)"),
        (18, "Savings_Check", "Flag measures that exceed ~20–30% of baseline (2nd-eyes)"),
    ]
    for r, a, b in rows:
        ws[f"A{r}"] = a
        ws[f"B{r}"] = b
    if payload.get("notes"):
        ws["A20"] = str(payload["notes"])


def _write_provenance(ws, prov: dict[str, Any]) -> None:
    ws["A1"] = "Model provenance"
    ws["A3"] = "key"
    ws["B3"] = "value"
    keys = [
        "idf_path",
        "idf_mtime",
        "epw",
        "g14_pass",
        "nmbe",
        "cvrmse",
        "cascade_ts",
        "building_id",
        "notes",
    ]
    r = 4
    for key in keys:
        if key in prov:
            ws[f"A{r}"] = key
            ws[f"B{r}"] = prov[key]
            r += 1
    for key, val in prov.items():
        if key in keys:
            continue
        ws[f"A{r}"] = key
        ws[f"B{r}"] = val
        r += 1


def _write_inputs(ws, inputs: list[dict[str, Any]] | dict[str, Any]) -> None:
    ws["A1"] = "Building Inputs (named ranges for Measures formulas)"
    ws["A2"] = (
        "Yellow / FITTED_FROM_EPLUS FLH rows are reverse-solved from E+ so "
        "formula × hours ≈ eplus — that is FITTED, not independent validation."
    )
    ws["A4"] = "name"
    ws["B4"] = "value"
    ws["C4"] = "provenance"
    ws["D4"] = "assumption_note"
    rows: list[dict[str, Any]]
    if isinstance(inputs, dict):
        rows = [
            {
                "name": k,
                "value": v.get("value", v) if isinstance(v, dict) else v,
                "provenance": (v.get("provenance") if isinstance(v, dict) else "VALUE"),
                "assumption_note": (v.get("note") if isinstance(v, dict) else ""),
            }
            for k, v in inputs.items()
        ]
    else:
        rows = list(inputs)
    for i, row in enumerate(rows, start=5):
        ws[f"A{i}"] = row.get("name")
        ws[f"B{i}"] = row.get("value")
        ws[f"C{i}"] = row.get("provenance") or row.get("hours_provenance") or "VALUE"
        ws[f"D{i}"] = row.get("assumption_note") or row.get("note") or ""


def _write_industry(ws, payload: dict[str, Any]) -> None:
    ws["A1"] = "Industry-method screening (independent — not fitted to EnergyPlus)"
    ws["A2"] = (
        "Formulas mirror ESCO VAV/chiller calculators. SCHED_ALIGN = fan cut + "
        "cool cut + OAD 0% recirculation."
    )
    industry = payload.get("industry_screening") or payload.get("industry") or {}
    ws["A4"] = "name"
    ws["B4"] = "value"
    ws["C4"] = "note"
    defaults = {
        "cfm_per_ton": (400, "ESCO books use metered Unit CFM when known"),
        "oa_frac": (0.2, "OA/SA fraction for mixed-air enthalpy"),
        "sched_current_weekly_h": (71.25, "Current AHU weekly hours"),
        "sched_future_weekly_h": (49.5, "Future opt-start weekly hours"),
        "sched_override_pad": (1.1, "Industry 10% allowance"),
        "warmup_cooldown_h": (450, "OAD 0% opportunity hours"),
        "SCHED_ALIGN_kwh": (
            industry.get("SCHED_ALIGN_kwh"),
            "fan + cool + OAD industry total",
        ),
    }
    r = 5
    seen = set()
    for name, (val, note) in defaults.items():
        if name in industry:
            val = industry[name]
        ws[f"A{r}"] = name
        ws[f"B{r}"] = val
        ws[f"C{r}"] = note
        seen.add(name)
        r += 1
    for name, val in industry.items():
        if name in seen:
            continue
        ws[f"A{r}"] = name
        ws[f"B{r}"] = val if not isinstance(val, dict) else val.get("value")
        ws[f"C{r}"] = val.get("note") if isinstance(val, dict) else ""
        r += 1


def _write_measures(
    ws,
    measures: list[dict[str, Any]],
    *,
    baseline_kwh: float | None,
    elec_rate: float,
) -> None:
    ws["A1"] = "Measure roster (building total)"
    ws["A2"] = (
        "Industry screening = 2nd set of eyes. FITTED wiring_echo_pct≈0 is not "
        "independent validation — use pct_diff_industry_vs_eplus."
    )
    headers = [
        "measure_id",
        "name",
        "eplus_source",
        "fitted_sheet_kwh",
        "eplus_kwh",
        "wiring_echo_pct",
        "industry_screen_kwh",
        "pct_diff_industry_vs_eplus",
        "annual_usd",
        "status",
        "hours_provenance",
        "pct_of_baseline",
        "savings_check",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(4, c, h)

    for i, m in enumerate(measures, start=5):
        fitted = m.get("fitted_sheet_kwh", m.get("fitted_kwh"))
        eplus = m.get("eplus_kwh")
        industry = m.get("industry_screen_kwh", m.get("industry_kwh"))
        prov = m.get("hours_provenance")
        status = m.get("status")
        if status is None:
            status = classify_measure_status(
                hours_provenance=prov,
                eplus_kwh=eplus,
                industry_kwh=industry,
                fitted_kwh=fitted,
                eplus_source=m.get("eplus_source"),
                sign_ok=m.get("sign_ok"),
            )
        if isinstance(status, MeasureHonestyStatus):
            status = status.value
        echo = m.get("wiring_echo_pct")
        if echo is None:
            echo = wiring_echo_pct(
                float(fitted) if fitted is not None else None,
                float(eplus) if eplus is not None else None,
            )
        ind_pct = m.get("pct_diff_industry_vs_eplus")
        if ind_pct is None:
            ind_pct = pct_diff_industry_vs_eplus(
                float(industry) if industry is not None else None,
                float(eplus) if eplus is not None else None,
            )
        annual = m.get("annual_usd")
        if annual is None and fitted is not None:
            annual = float(fitted) * elec_rate
        pct_base = m.get("pct_of_baseline")
        if pct_base is None and fitted is not None and baseline_kwh:
            pct_base = float(fitted) / float(baseline_kwh)
        check = m.get("savings_check")
        if check is None and pct_base is not None:
            check = "REVIEW" if pct_base > 0.30 else ("WATCH" if pct_base > 0.20 else "OK")

        values = [
            m.get("measure_id"),
            m.get("name"),
            m.get("eplus_source"),
            fitted,
            eplus,
            echo,
            industry,
            ind_pct,
            annual,
            status,
            prov,
            pct_base,
            check,
        ]
        for c, v in enumerate(values, 1):
            ws.cell(i, c, v)


def _write_demand(ws, demand: dict[str, Any]) -> None:
    ws["A1"] = "Demand (kW) — twin / July pair"
    ws["A3"] = "metric"
    ws["B3"] = "kW"
    mapping = [
        ("july_weekday_kw", demand.get("july_weekday_kw", demand.get("weekday_kw"))),
        ("july_weekend_kw", demand.get("july_weekend_kw", demand.get("weekend_kw"))),
        ("loadshed_kw", demand.get("loadshed_kw", demand.get("load_shed_kw"))),
    ]
    for i, (name, val) in enumerate(mapping, start=4):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = val


def _write_twin_calibrate(ws, twin_cal: dict[str, Any]) -> None:
    ws["A1"] = "Twin calibration"
    ws["A3"] = "key"
    ws["B3"] = "value"
    if isinstance(twin_cal, dict):
        for i, (k, v) in enumerate(twin_cal.items(), start=4):
            ws[f"A{i}"] = k
            ws[f"B{i}"] = v if not isinstance(v, (list, dict)) else str(v)
