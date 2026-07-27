"""Engineering Findings Excel notebook — punchlist-first (openpyxl)."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from open_fdd.reporting.models import ReportArtifacts

REQUIRED_SHEETS = (
    "Punchlist",
    "Findings",
    "FAULT_Inventory",
    "Overview_Charts",
    "Suppressed_DQ",
    "Quality_Gate",
)


def render_findings_xlsx(
    artifacts: ReportArtifacts,
    out_path: Path | str,
    *,
    embed_images: bool = True,
) -> Path:
    """Write punchlist-first Eng Findings workbook. Returns path."""
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Engineering Findings Excel. "
            "Install with: pip install '.[engineering-report]' or openpyxl>=3.1"
        ) from exc

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    def _hdr(ws, row: int = 1) -> None:
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font

    # --- Punchlist (first) ---
    punch = wb.active
    punch.title = "Punchlist"
    punch.append(["done", "item", "finding_id", "equipment", "priority"])
    _hdr(punch)
    # Expand checklist with finding provenance when possible
    checklist = list(artifacts.field_checklist or [])
    included = [f for f in artifacts.findings if f.include_in_report]
    if checklist:
        for item in checklist:
            fid = ""
            equip = ""
            pri = ""
            for f in included:
                if item in (f.field_verification or []):
                    fid = f.finding_id
                    equip = ", ".join(f.equipment_ids[:3])
                    pri = f.priority
                    break
            punch.append(["[ ]", item, fid, equip, pri])
    else:
        for f in included:
            for item in f.field_verification or []:
                punch.append(["[ ]", item, f.finding_id, ", ".join(f.equipment_ids[:3]), f.priority])
    punch.column_dimensions["B"].width = 72
    punch.column_dimensions["C"].width = 12
    punch.column_dimensions["D"].width = 24

    # --- Findings ---
    find_ws = wb.create_sheet("Findings")
    find_ws.append(
        [
            "finding_id",
            "priority",
            "title",
            "classification",
            "systems",
            "equipment_ids",
            "rule_ids",
            "why_it_matters",
            "observed_behavior",
            "day_zoom_path",
            "day_zoom_error",
            "engineer_note",
        ]
    )
    _hdr(find_ws)
    for i, f in enumerate(included, start=2):
        note = (f.engineer_override or {}).get("note", "")
        # Relative basename for agent-friendly cells; PNG still embeds from absolute (BUG-039)
        zoom_cell = ""
        if f.day_zoom_path:
            try:
                zoom_cell = Path(str(f.day_zoom_path)).name
            except Exception:
                zoom_cell = str(f.day_zoom_path)
        find_ws.append(
            [
                f.finding_id,
                f.priority,
                f.title,
                f.effective_classification.value,
                ", ".join(f.systems or []),
                ", ".join(f.equipment_ids or []),
                ", ".join(f.rule_ids or []),
                f.why_it_matters,
                f.observed_behavior,
                zoom_cell,
                f.day_zoom_skip_reason,
                note,
            ]
        )
        if embed_images and f.day_zoom_path and Path(f.day_zoom_path).is_file():
            try:
                img = XLImage(f.day_zoom_path)
                img.width = min(img.width, 480)
                img.height = min(img.height, 280)
                # Place below data block starting at column N
                find_ws.add_image(img, f"N{i}")
            except Exception:
                pass
    find_ws.column_dimensions["C"].width = 40
    find_ws.column_dimensions["H"].width = 36
    find_ws.column_dimensions["I"].width = 36

    # --- FAULT inventory ---
    inv_ws = wb.create_sheet("FAULT_Inventory")
    inv = artifacts.fault_inventory or {}
    inv_ws.append(
        [
            "equipment",
            "equipment_prefix",
            "system",
            "rule_id",
            "fault_samples",
            "fault_hours",
            "in_priority",
            "suppressed_reason",
            "candidate_key",
        ]
    )
    _hdr(inv_ws)
    for row in inv.get("rows") or []:
        inv_ws.append(
            [
                row.get("equipment"),
                row.get("equipment_prefix"),
                row.get("system"),
                row.get("rule_id"),
                row.get("fault_samples"),
                row.get("fault_hours"),
                row.get("in_priority"),
                row.get("suppressed_reason"),
                row.get("candidate_key"),
            ]
        )
    inv_ws.append([])
    inv_ws.append(
        [
            "rollup",
            "n_faults",
            "n_in_priority_findings",
            "n_candidates_in_priority",
            "n_orphans",
        ]
    )
    inv_ws.append(
        [
            "summary",
            inv.get("n_faults"),
            inv.get("n_priority_findings", inv.get("n_in_priority")),
            inv.get("n_candidates_in_priority"),
            inv.get("n_orphans"),
        ]
    )

    # --- Overview charts (image embeds) ---
    ov = wb.create_sheet("Overview_Charts")
    ov["A1"] = "Overview / summary chart paths"
    ov["A1"].font = Font(bold=True)
    ov.append(["name", "path", "embedded"])
    _hdr(ov, row=2)
    r = 3
    img_row = 3
    # Prefer overview_charts; fall back to charts — dedupe by name (BUG-038)
    seen_names: set[str] = set()
    chart_metas: list[dict[str, Any]] = []
    for meta in list(artifacts.overview_charts or []) + list(artifacts.charts or []):
        if not isinstance(meta, dict):
            continue
        name = str(meta.get("name") or meta.get("title") or "").strip()
        if not name or name in seen_names:
            continue
        seen_names.add(name)
        chart_metas.append(meta)
    for meta in chart_metas:
        name = meta.get("name") or meta.get("title") or ""
        path = meta.get("path")
        # Prefer relative path for agent-friendly cells (BUG-039)
        path_cell = path
        if path:
            try:
                path_cell = str(Path(str(path)).name)
            except Exception:
                path_cell = path
        embedded = False
        ov[f"A{r}"] = name
        ov[f"B{r}"] = path_cell
        if (
            embed_images
            and path
            and Path(str(path)).is_file()
            and str(path).lower().endswith((".png", ".jpg", ".jpeg"))
        ):
            try:
                img = XLImage(str(path))
                img.width = min(getattr(img, "width", 640) or 640, 640)
                img.height = min(getattr(img, "height", 360) or 360, 360)
                ov.add_image(img, f"D{img_row}")
                img_row += 18
                embedded = True
            except Exception:
                embedded = False
        ov[f"C{r}"] = embedded
        r += 1
    ov.column_dimensions["A"].width = 28
    ov.column_dimensions["B"].width = 64

    # --- Suppressed / DQ ---
    dq = wb.create_sheet("Suppressed_DQ")
    dq.append(["kind", "candidate_key", "classification", "score", "equipment_id", "rule_id", "reasons"])
    _hdr(dq)
    for row in artifacts.suppressed or []:
        reasons = row.get("reasons") or []
        dq.append(
            [
                "suppressed",
                row.get("candidate_key"),
                row.get("classification"),
                row.get("score"),
                row.get("equipment_id"),
                row.get("rule_id"),
                "; ".join(str(x) for x in reasons),
            ]
        )
    for row in artifacts.data_quality or []:
        reasons = row.get("reasons") or []
        dq.append(
            [
                "data_quality",
                row.get("candidate_key"),
                row.get("classification"),
                row.get("score"),
                row.get("equipment_id"),
                row.get("rule_id"),
                "; ".join(str(x) for x in reasons),
            ]
        )
    dq.column_dimensions["G"].width = 48

    # --- Quality gate ---
    qg = wb.create_sheet("Quality_Gate")
    qg["A1"] = "Building"
    qg["B1"] = artifacts.building
    qg["A2"] = "Analysis period"
    qg["B2"] = artifacts.analysis_period
    qg["A3"] = "Generated"
    qg["B3"] = artifacts.generated_at
    qg["A4"] = "Gate OK"
    qg["B4"] = (artifacts.quality_gate or {}).get("ok")
    qg["A5"] = "Errors"
    qg["B5"] = "; ".join((artifacts.quality_gate or {}).get("errors") or [])
    qg["A6"] = "Warnings"
    qg["B6"] = "; ".join((artifacts.quality_gate or {}).get("warnings") or [])
    qg["A8"] = "Metrics"
    qg["A8"].font = Font(bold=True)
    r = 9
    for k, v in (artifacts.metrics or {}).items():
        qg[f"A{r}"] = k
        qg[f"B{r}"] = json_safe(v)
        r += 1
    qg.column_dimensions["A"].width = 28
    qg.column_dimensions["B"].width = 72

    wb.save(out_path)
    return out_path


def json_safe(v: Any) -> Any:
    if isinstance(v, (str, int, float, bool)) or v is None:
        return v
    try:
        import json

        return json.dumps(v)
    except Exception:
        return str(v)
