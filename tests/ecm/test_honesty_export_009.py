"""BUG-OFDD-ECM-009 — honesty export sheets."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from open_fdd.ecm_engineering import ECMJob
from open_fdd.ecm_engineering.honesty_export import HONESTY_SHEETS


def test_honesty_export_contents_and_demand(tmp_path: Path) -> None:
    out = tmp_path / "honesty.xlsx"
    job = ECMJob("Liberty dual AHU", path=tmp_path / "toolkit.xlsx")
    job.attach_twin_compare(
        {
            "provenance": {
                "idf_path": "/models/b100.idf",
                "epw": "TMY3",
                "g14_pass": True,
                "nmbe": 0.02,
                "cvrmse": 0.08,
                "cascade_ts": "2026-07-29T12:00:00Z",
            },
            "inputs": [
                {
                    "name": "lockout_hours",
                    "value": 612.2,
                    "provenance": "FITTED_FROM_EPLUS",
                    "assumption_note": "eplus / plant_kw",
                },
                {
                    "name": "loadshed_hours",
                    "value": 2,
                    "provenance": "SCREENING_ASSUMPTION",
                },
            ],
            "industry_screening": {"SCHED_ALIGN_kwh": 72000.0},
            "measures": [
                {
                    "measure_id": "ECM-CHILLER-LOCKOUT",
                    "name": "Chiller OAT lockout",
                    "eplus_source": "cascade",
                    "fitted_sheet_kwh": 101580.56,
                    "eplus_kwh": 101580.56,
                    "industry_screen_kwh": 95000.0,
                    "hours_provenance": "FITTED_FROM_EPLUS",
                },
                {
                    "measure_id": "ECM-LOAD-SHED-DR",
                    "name": "Load shed",
                    "eplus_source": "July MCP pair",
                    "fitted_sheet_kwh": 39.0,
                    "eplus_kwh": 69.0,
                    "hours_provenance": "SCREENING_ASSUMPTION",
                },
            ],
            "baseline_kwh": 1_500_000.0,
            "demand": {
                "july_weekday_kw": 420.0,
                "july_weekend_kw": 280.0,
                "loadshed_kw": 365.0,
            },
            "twin_calibrate": {"g14": "PASS"},
        }
    )
    path = job.save(out)
    assert path == out.resolve() or path == out
    wb = load_workbook(out)
    for name in HONESTY_SHEETS:
        assert name in wb.sheetnames
    assert "Demand" in wb.sheetnames
    assert "Twin_Calibrate" in wb.sheetnames
    assert "Cover" not in wb.sheetnames
    assert "Formula_Trace" not in wb.sheetnames

    measures = wb["Measures"]
    headers = [measures.cell(4, c).value for c in range(1, 14)]
    assert "wiring_echo_pct" in headers
    assert "pct_diff_industry_vs_eplus" in headers
    assert "status" in headers
    statuses = [measures.cell(r, 10).value for r in range(5, 7)]
    assert "FITTED" in statuses
    assert "BALLPARK" in statuses

    demand = wb["Demand"]
    assert demand["A4"].value == "july_weekday_kw"
    assert demand["B4"].value == 420.0
